"""
Stowage Plan Optimizer — OR-Tools CP-SAT
Reads MongoDB, generates 5 alternative stowage plans, exports Excel + JSON.
Usage: python solver.py <voyage_id>
       python solver.py          (lists available voyages)
"""

import sys
import os
import math
import json
import re
from datetime import datetime
from collections import defaultdict

from dotenv import load_dotenv
from pymongo import MongoClient
from bson import ObjectId
from ortools.sat.python import cp_model
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# ── Constants ──────────────────────────────────────────────────────────────────

# Higher index = deeper / less accessible (D is bottom, FC is deck)
LEVEL_DEPTH = {'FC': 0, 'UPD': 1, 'A': 2, 'B': 3, 'C': 4, 'D': 5}

CARGO_TEMP_RANGES = {
    'BANANAS':          (12,    14),
    'ORGANIC_BANANAS':  (12,    14),
    'PLANTAINS':        (12,    14),
    'FROZEN_FISH':      (-25,  -18),
    'OTHER_FROZEN':     (-25,  -18),
    'TABLE_GRAPES':     (-0.5,  0.5),
    'BERRIES':          (-0.5,  1.0),
    'CHERRIES':         (-0.5,  0.5),
    'BLUEBERRIES':      (-0.5,  1.0),
    'CITRUS':           (4,    10),
    'AVOCADOS':         (5,     8),
    'PINEAPPLES':       (7,    10),
    'PAPAYA':           (7,    10),
    'MANGOES':          (7,    10),
    'KIWIS':            (-0.5,  1),
    'APPLES':           (-0.5,  4),
    'PEARS':            (-0.5,  4),
    'PLUMS':            (0.0,   2.0),
    'PEACHES':          (-0.5,  2),
    'OTHER_CHILLED':    (2,     6),
}

# Matches lib/constants/pod-colors.ts (without '#')
POD_EXCEL_COLORS = {
    'NLVLI': 'F97316', 'NLRTM': 'F59E0B',
    'GBPME': '3B82F6', 'GBSOU': '60A5FA', 'GBFXT': '93C5FD', 'GBDVR': '2563EB',
    'BEANR': 'EAB308',
    'DEHAM': '6B7280', 'DEBHV': '94A3B8',
    'FRLEH': '8B5CF6', 'FRRAD': '7C3AED',
    'USLAX': 'EF4444', 'USNYC': 'F87171', 'USORF': 'FCA5A5',
    'USSAV': 'DC2626', 'USWIL': 'B91C1C', 'USMIA': 'E11D48', 'USBAL': 'BE123C',
    'ESBCN': 'D97706', 'ESVLC': 'B45309',
    'ITGOA': '10B981', 'ITLIV': '059669',
}
_FALLBACK_COLORS = ['06B6D4', 'EC4899', 'A78BFA', '34D399',
                    'FB923C', 'F472B6', '38BDF8', 'A3E635']

SOLUTION_CONFIGS = [
    {'label': 'Balanced',        'W_OV': 10000, 'W_BAL': 1000, 'W_CMP': 100,  'W_POD': 10,  'W_UTL': 1},
    {'label': 'Max Balance',     'W_OV': 10000, 'W_BAL': 5000, 'W_CMP': 50,   'W_POD': 10,  'W_UTL': 1},
    {'label': 'Max Compactness', 'W_OV': 10000, 'W_BAL': 500,  'W_CMP': 500,  'W_POD': 10,  'W_UTL': 1},
    {'label': 'POD-Friendly',    'W_OV': 10000, 'W_BAL': 800,  'W_CMP': 100,  'W_POD': 200, 'W_UTL': 1},
    {'label': 'Max Utilization', 'W_OV': 10000, 'W_BAL': 500,  'W_CMP': 100,  'W_POD': 10,  'W_UTL': 100},
]

STATUS_NAMES = {
    cp_model.OPTIMAL:    'OPTIMAL',
    cp_model.FEASIBLE:   'FEASIBLE',
    cp_model.INFEASIBLE: 'INFEASIBLE',
    cp_model.UNKNOWN:    'UNKNOWN (time limit)',
}

# ── Helpers ────────────────────────────────────────────────────────────────────

def parse_section_id(sid):
    """'2UPD' → (2, 'UPD'),  '1A' → (1, 'A'),  '3FC' → (3, 'FC')"""
    m = re.match(r'^(\d+)([A-Za-z]+)$', sid)
    if m:
        return int(m.group(1)), m.group(2).upper()
    return 0, 'A'

def get_temp_range(cargo_type):
    return CARGO_TEMP_RANGES.get(cargo_type, (0, 4))

def get_zone_temperature(zone):
    t = zone.get('currentTemperature')
    if t is not None:
        return float(t)
    tr = zone.get('temperatureRange') or {}
    mn, mx = tr.get('min'), tr.get('max')
    if mn is not None and mx is not None:
        return (float(mn) + float(mx)) / 2.0
    return None

def get_pod_excel_color(port_code):
    if port_code in POD_EXCEL_COLORS:
        return POD_EXCEL_COLORS[port_code]
    h = 0
    for c in (port_code or ''):
        h = ord(c) + ((h << 5) - h)
    return _FALLBACK_COLORS[abs(h) % len(_FALLBACK_COLORS)]

def is_temp_compatible(cargo, section):
    t = section['temperature']
    if t is None:
        return True
    return cargo['tempMin'] <= t <= cargo['tempMax']

def temp_ranges_overlap(cargo_a, cargo_b):
    return cargo_a['tempMax'] >= cargo_b['tempMin'] and \
           cargo_b['tempMax'] >= cargo_a['tempMin']

# ── MongoDB data loading ───────────────────────────────────────────────────────

def load_voyage(db, voyage_id):
    return db.voyages.find_one({'_id': ObjectId(voyage_id)})

def load_vessel(db, vessel_id):
    return db.vessels.find_one({'_id': vessel_id})

def load_bookings(db, voyage_id):
    return list(db.bookings.find({
        'voyageId': ObjectId(voyage_id),
        'status': {'$in': ['CONFIRMED', 'PARTIAL']},
    }))

def load_forecasts(db, voyage_id):
    return list(db.spaceforecasts.find({
        'voyageId': ObjectId(voyage_id),
        'source': {'$in': ['PLANNER_ENTRY', 'SHIPPER_PORTAL', 'CONTRACT_DEFAULT']},
        'planImpact': 'INCORPORATED',
        'estimatedPallets': {'$gt': 0},
    }))

def build_port_seq_map(port_calls):
    return {pc['portCode']: pc['sequence'] for pc in port_calls}

def build_sections(vessel):
    sections = []
    for zone in vessel.get('temperatureZones', []):
        zone_id = zone.get('zoneId', '')
        temp = get_zone_temperature(zone)
        for cs in zone.get('coolingSections', []):
            sid = cs.get('sectionId', '')
            sqm = float(cs.get('sqm') or 0)
            dsf = float(cs.get('designStowageFactor') or 1.32)
            capacity = math.floor(sqm * dsf)
            hold_num, level = parse_section_id(sid)
            sections.append({
                'sectionId':   sid,
                'zoneId':      zone_id,
                'holdNumber':  hold_num,
                'level':       level,
                'levelDepth':  LEVEL_DEPTH.get(level, 2),
                'sqm':         sqm,
                'capacity':    capacity,
                'temperature': temp,
            })
    return sections

def build_cargo_items(bookings, forecasts, port_seq_map):
    items = []
    for b in bookings:
        pallets = int(b.get('confirmedQuantity') or 0)
        if pallets <= 0:
            continue
        pol = b.get('pol', {}).get('portCode', '')
        pod = b.get('pod', {}).get('portCode', '')
        tmin, tmax = get_temp_range(b.get('cargoType', ''))
        items.append({
            'id':            str(b['_id']),
            'bookingId':     str(b['_id']),
            'bookingNumber': b.get('bookingNumber', ''),
            'cargoType':     b.get('cargoType', ''),
            'pallets':       pallets,
            'polPortCode':   pol,
            'podPortCode':   pod,
            'polSeq':        port_seq_map.get(pol, 0),
            'podSeq':        port_seq_map.get(pod, 99),
            'tempMin':       tmin,
            'tempMax':       tmax,
            'shipperName':   b.get('shipper', {}).get('name', ''),
            'consigneeName': b.get('consignee', {}).get('name', ''),
            'confidence':    'CONFIRMED',
        })
    for f in forecasts:
        pallets = int(f.get('estimatedPallets') or 0)
        if pallets <= 0:
            continue
        pol = f.get('polPortCode', '')
        pod = f.get('podPortCode', '')
        tmin, tmax = get_temp_range(f.get('cargoType', ''))
        items.append({
            'id':            f'FORECAST-{f["_id"]}',
            'bookingId':     f'FORECAST-{f["_id"]}',
            'bookingNumber': f'EST-{f.get("forecastNumber", "")}',
            'cargoType':     f.get('cargoType', ''),
            'pallets':       pallets,
            'polPortCode':   pol,
            'podPortCode':   pod,
            'polSeq':        port_seq_map.get(pol, 0),
            'podSeq':        port_seq_map.get(pod, 99),
            'tempMin':       tmin,
            'tempMax':       tmax,
            'shipperName':   f.get('shipperName', ''),
            'consigneeName': f.get('consigneeName', ''),
            'confidence':    'ESTIMATED',
        })
    return items

def print_loading_summary(voyage, vessel, bookings, forecasts, sections):
    port_calls = sorted(voyage.get('portCalls', []), key=lambda p: p['sequence'])
    rotation = ' -> '.join(f'{p["portCode"]}(seq={p["sequence"]})' for p in port_calls)
    total_pal = (sum(int(b.get('confirmedQuantity') or 0) for b in bookings) +
                 sum(int(f.get('estimatedPallets') or 0) for f in forecasts))
    total_cap = sum(s['capacity'] for s in sections)
    util = total_pal / total_cap * 100 if total_cap else 0
    holds = len(set(s['holdNumber'] for s in sections))

    print(f"Voyage: {voyage.get('voyageNumber', '?')} | Vessel: {vessel.get('name', '?')}")
    print(f"Port rotation: {rotation}")
    print(f"Bookings: {len(bookings)} | Forecasts: {len(forecasts)}")
    print(f"Total pallets: {total_pal} | Vessel capacity: {total_cap} ({util:.1f}%)")
    print(f"Sections: {len(sections)} across {holds} holds")
    print(f"Temperature zones: {len(vessel.get('temperatureZones', []))}")
    for zone in vessel.get('temperatureZones', []):
        zid  = zone.get('zoneId', '?')
        temp = get_zone_temperature(zone)
        secs = [cs.get('sectionId', '') for cs in zone.get('coolingSections', [])]
        temp_str = f'{temp}°C' if temp is not None else 'unset'
        print(f"  {zid}: {temp_str}: {', '.join(secs)}")

# ── CP-SAT solver ──────────────────────────────────────────────────────────────

def solve(cargo_items, sections, config, time_limit=30):
    model  = cp_model.CpModel()
    n_c    = len(cargo_items)
    n_s    = len(sections)
    compat = {(i, j) for i in range(n_c) for j in range(n_s)
              if is_temp_compatible(cargo_items[i], sections[j])}

    # Decision vars: x[i,j] = pallets of cargo i in section j
    x = {}
    for i in range(n_c):
        for j in range(n_s):
            if (i, j) in compat:
                ub = min(cargo_items[i]['pallets'], sections[j]['capacity'])
                x[i, j] = model.NewIntVar(0, ub, f'x{i}_{j}')
            else:
                x[i, j] = model.NewIntVar(0, 0, f'z{i}_{j}')

    total_cap = sum(s['capacity'] for s in sections)
    total_pal = sum(c['pallets'] for c in cargo_items)

    # ── Constraint 1: SUPPLY — each cargo fully placed ──────────────────────
    for i in range(n_c):
        model.Add(sum(x[i, j] for j in range(n_s)) == cargo_items[i]['pallets'])

    # ── Constraint 2: CAPACITY — sections not overfilled ────────────────────
    for j in range(n_s):
        model.Add(sum(x[i, j] for i in range(n_c)) <= sections[j]['capacity'])

    # ── Constraints 3+4: POL + POD MONOTONICITY ─────────────────────────────
    # Group section indices by hold
    holds_map = defaultdict(list)
    for j, sec in enumerate(sections):
        holds_map[sec['holdNumber']].append(j)

    _bv = [0]  # mutable counter for unique bool-var names

    for hold_num, h_secs in holds_map.items():
        # Sort by levelDepth ascending (FC=0 = shallowest, D=5 = deepest)
        sorted_secs = sorted(h_secs, key=lambda j: sections[j]['levelDepth'])

        for hi_idx, j_hi in enumerate(sorted_secs):          # j_hi = shallower
            for j_lo in sorted_secs[hi_idx + 1:]:            # j_lo = deeper

                for i1 in range(n_c):
                    for i2 in range(i1 + 1, n_c):

                        # POL monotonicity:
                        # Overstow check counts early-in-j_lo + late-in-j_hi as a
                        # violation (early cargo in the higher-depth section blocks
                        # late cargo from reaching the lower-depth section at a later
                        # POL). Block that specific pattern.
                        pol1, pol2 = cargo_items[i1]['polSeq'], cargo_items[i2]['polSeq']
                        if pol1 != pol2:
                            i_early = i1 if pol1 < pol2 else i2
                            i_late  = i2 if pol1 < pol2 else i1
                            if (i_early, j_lo) in compat and (i_late, j_hi) in compat:
                                b = model.NewBoolVar(f'b_pol_{_bv[0]}'); _bv[0] += 1
                                cap_lo = sections[j_lo]['capacity']
                                cap_hi = sections[j_hi]['capacity']
                                model.Add(x[i_early, j_lo] <= cap_lo * b)
                                model.Add(x[i_early, j_lo] >= b)
                                model.Add(x[i_late, j_hi] <= cap_hi * (1 - b))

                        # POD monotonicity:
                        # Early-discharge cargo (lower podSeq, first port to unload)
                        # must be accessible: not buried under late-discharge cargo.
                        # Block early-discharge in j_lo + late-discharge in j_hi.
                        pod1, pod2 = cargo_items[i1]['podSeq'], cargo_items[i2]['podSeq']
                        if pod1 != pod2:
                            i_early_d = i1 if pod1 < pod2 else i2
                            i_late_d  = i2 if pod1 < pod2 else i1
                            if (i_early_d, j_lo) in compat and (i_late_d, j_hi) in compat:
                                b = model.NewBoolVar(f'b_pod_{_bv[0]}'); _bv[0] += 1
                                cap_lo = sections[j_lo]['capacity']
                                cap_hi = sections[j_hi]['capacity']
                                model.Add(x[i_early_d, j_lo] <= cap_lo * b)
                                model.Add(x[i_early_d, j_lo] >= b)
                                model.Add(x[i_late_d, j_hi] <= cap_hi * (1 - b))

    # ── Constraint 5: TEMPERATURE ZONE GROUPING ────────────────────────────
    # Two cargo types with non-overlapping temperature ranges must not both
    # be placed in the same cooling zone (zoneId).
    zones_sec = defaultdict(list)
    for j, sec in enumerate(sections):
        zones_sec[sec['zoneId']].append(j)

    incompat_pairs = [
        (i1, i2)
        for i1 in range(n_c) for i2 in range(i1 + 1, n_c)
        if not temp_ranges_overlap(cargo_items[i1], cargo_items[i2])
    ]

    for i1, i2 in incompat_pairs:
        for zone_id, z_secs in zones_sec.items():
            b = model.NewBoolVar(f'tz_{i1}_{i2}_{zone_id}')
            model.Add(sum(x[i1, j] for j in z_secs) == 0).OnlyEnforceIf(b.Not())
            model.Add(sum(x[i2, j] for j in z_secs) == 0).OnlyEnforceIf(b)

    # ── Objective ──────────────────────────────────────────────────────────
    W_BAL = config['W_BAL']
    W_CMP = config['W_CMP']
    W_POD = config['W_POD']
    W_UTL = config['W_UTL']

    obj = []

    # Balance: |fwd_pallets − aft_pallets|
    fwd_js = [j for j in range(n_s) if sections[j]['holdNumber'] in {1, 2}]
    aft_js = [j for j in range(n_s) if sections[j]['holdNumber'] in {3, 4}]
    if fwd_js and aft_js:
        fwd_v = model.NewIntVar(0, total_cap, 'fwd')
        aft_v = model.NewIntVar(0, total_cap, 'aft')
        model.Add(fwd_v == sum(x[i, j] for i in range(n_c) for j in fwd_js))
        model.Add(aft_v == sum(x[i, j] for i in range(n_c) for j in aft_js))
        bal = model.NewIntVar(0, total_cap, 'bal')
        model.AddAbsEquality(bal, fwd_v - aft_v)
        obj.append(W_BAL * bal)

    # Compactness: penalize pallets in shallow sections when deeper sections empty
    for hold_num, h_secs in holds_map.items():
        by_depth = sorted(h_secs, key=lambda j: sections[j]['levelDepth'], reverse=True)
        for d_idx, j_deep in enumerate(by_depth[:-1]):
            shallow_above = by_depth[d_idx + 1:]
            deep_sum_expr = sum(x[i, j_deep] for i in range(n_c))
            shallow_sum_expr = sum(x[i, j] for j in shallow_above for i in range(n_c))
            deep_v = model.NewIntVar(0, sections[j_deep]['capacity'], f'dv{j_deep}')
            model.Add(deep_v == deep_sum_expr)
            b_empty = model.NewBoolVar(f'de{j_deep}')
            model.Add(deep_v == 0).OnlyEnforceIf(b_empty)
            model.Add(deep_v >= 1).OnlyEnforceIf(b_empty.Not())
            pen = model.NewIntVar(0, total_pal, f'cp{j_deep}')
            model.Add(pen == shallow_sum_expr).OnlyEnforceIf(b_empty)
            model.Add(pen == 0).OnlyEnforceIf(b_empty.Not())
            obj.append(W_CMP * pen)

    # POD distribution: penalize same-POD concentration in one hold
    pod_codes = list(set(c['podPortCode'] for c in cargo_items))
    for pod in pod_codes:
        p_idxs = [i for i, c in enumerate(cargo_items) if c['podPortCode'] == pod]
        if len(p_idxs) < 2:
            continue
        pod_total = sum(cargo_items[i]['pallets'] for i in p_idxs)
        half = pod_total // 2
        for hold_num in range(1, 5):
            h_js = [j for j in range(n_s) if sections[j]['holdNumber'] == hold_num]
            if not h_js:
                continue
            pod_hold_v = model.NewIntVar(0, pod_total, f'ph_{pod}_{hold_num}')
            model.Add(pod_hold_v == sum(x[i, j] for i in p_idxs for j in h_js))
            excess = model.NewIntVar(-pod_total, pod_total, f'ex_{pod}_{hold_num}')
            model.Add(excess == pod_hold_v - half)
            conc = model.NewIntVar(0, pod_total, f'cn_{pod}_{hold_num}')
            model.AddMaxEquality(conc, [excess, model.NewConstant(0)])
            obj.append(W_POD * conc)

    # Utilization: maximize total placed pallets
    placed = sum(x[i, j] for i in range(n_c) for j in range(n_s))
    obj.append(-W_UTL * placed)

    model.Minimize(sum(obj))

    solver = cp_model.CpSolver()
    solver.parameters.max_time_in_seconds = time_limit
    solver.parameters.num_search_workers   = 4
    status = solver.Solve(model)

    if status not in (cp_model.OPTIMAL, cp_model.FEASIBLE):
        return None, status, solver

    assignments = [
        {'cargoIdx': i, 'sectionIdx': j, 'quantity': solver.Value(x[i, j])}
        for i in range(n_c) for j in range(n_s)
        if solver.Value(x[i, j]) > 0
    ]
    return assignments, status, solver

# ── Metrics ────────────────────────────────────────────────────────────────────

def compute_metrics(assignments, cargo_items, sections):
    total_pal = sum(c['pallets'] for c in cargo_items)
    placed    = sum(a['quantity'] for a in assignments)

    fwd = sum(a['quantity'] for a in assignments
              if sections[a['sectionIdx']]['holdNumber'] in {1, 2})
    aft = sum(a['quantity'] for a in assignments
              if sections[a['sectionIdx']]['holdNumber'] in {3, 4})
    balance_dev = abs(fwd - aft)

    # Overstow: late-POL cargo sitting below early-POL cargo in same hold
    overstow = 0
    by_hold = defaultdict(list)
    for a in assignments:
        sec = sections[a['sectionIdx']]
        by_hold[sec['holdNumber']].append({
            'depth':    sec['levelDepth'],
            'polSeq':   cargo_items[a['cargoIdx']]['polSeq'],
            'qty':      a['quantity'],
        })
    for h_assgns in by_hold.values():
        for a1 in h_assgns:
            for a2 in h_assgns:
                if a1['polSeq'] < a2['polSeq'] and a2['depth'] < a1['depth']:
                    overstow += min(a1['qty'], a2['qty'])

    lower = {'C', 'D'}
    lower_cap  = sum(s['capacity'] for s in sections if s['level'] in lower)
    lower_used = sum(a['quantity'] for a in assignments
                     if sections[a['sectionIdx']]['level'] in lower)
    compact_pct = lower_used / lower_cap * 100 if lower_cap else 0
    secs_used   = len(set(a['sectionIdx'] for a in assignments))

    return {
        'placedPallets':   placed,
        'totalPallets':    total_pal,
        'placedPct':       placed / total_pal * 100 if total_pal else 0,
        'overstowViolations': overstow,
        'balanceDev':      balance_dev,
        'compactnessPct':  compact_pct,
        'sectionsUsed':    secs_used,
        'totalSections':   len(sections),
    }

# ── Excel export ───────────────────────────────────────────────────────────────

def export_excel(solutions, voyage_number, out_dir):
    os.makedirs(out_dir, exist_ok=True)
    ts   = datetime.now().strftime('%Y%m%d_%H%M%S')
    path = os.path.join(out_dir, f'plan_{voyage_number}_{ts}.xlsx')
    wb   = Workbook()

    # Summary sheet
    ws = wb.active
    ws.title = 'Summary'
    hdrs = ['#', 'Label', 'Status', 'Placed', 'Total', 'Placed%',
            'Overstow', 'Balance Dev', 'Compact%', 'Sections']
    ws.append(hdrs)
    for cell in ws[1]:
        cell.font      = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    for sol in solutions:
        m = sol.get('metrics', {})
        ws.append([
            sol['solutionIndex'], sol['label'], sol['status'],
            m.get('placedPallets', ''), m.get('totalPallets', ''),
            f"{m.get('placedPct', 0):.1f}%",
            m.get('overstowViolations', ''),
            m.get('balanceDev', ''),
            f"{m.get('compactnessPct', 0):.1f}%",
            f"{m.get('sectionsUsed', '')}/{m.get('totalSections', '')}",
        ])
    ws.freeze_panes = 'A2'
    for c in range(1, len(hdrs) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 16

    # Per-solution sheets
    col_hdrs = ['Hold', 'Section', 'Level', 'Capacity', 'Cargo Type',
                'Shipper', 'Consignee', 'POL', 'POD', 'Pallets', 'Confidence']
    for sol in solutions:
        ws2 = wb.create_sheet(title=f"Plan {sol['solutionIndex']}")
        ws2.append(col_hdrs)
        for cell in ws2[1]:
            cell.font      = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        ws2.freeze_panes = 'A2'
        for pos in sorted(sol.get('cargoPositions', []),
                          key=lambda p: (p['holdNumber'], p['level'])):
            ws2.append([
                pos['holdNumber'], pos['sectionId'], pos['level'],
                pos['capacity'],   pos['cargoType'],
                pos['shipperName'], pos['consigneeName'],
                pos['polPortCode'], pos['podPortCode'],
                pos['quantity'],   pos['confidence'],
            ])
            color = get_pod_excel_color(pos.get('podPortCode', ''))
            fill  = PatternFill(start_color=color, end_color=color, fill_type='solid')
            for cell in ws2[ws2.max_row]:
                cell.fill = fill
        for c in range(1, len(col_hdrs) + 1):
            ws2.column_dimensions[get_column_letter(c)].width = 16

    wb.save(path)
    return path

# ── JSON export ────────────────────────────────────────────────────────────────

def export_json(voyage_id, solutions, voyage_number, out_dir):
    os.makedirs(out_dir, exist_ok=True)
    ts   = datetime.now().strftime('%Y%m%d_%H%M%S')
    path = os.path.join(out_dir, f'plan_{voyage_number}_{ts}.json')
    out  = {
        'voyageId':     voyage_id,
        'voyageNumber': voyage_number,
        'generatedAt':  datetime.utcnow().isoformat() + 'Z',
        'solutions':    [],
    }
    for sol in solutions:
        out['solutions'].append({
            'solutionIndex': sol['solutionIndex'],
            'label':         sol['label'],
            'status':        sol['status'],
            'metrics':       sol.get('metrics', {}),
            'cargoPositions': [
                {
                    'bookingId':     p['bookingId'],
                    'sectionId':     p['sectionId'],
                    'holdNumber':    p['holdNumber'],
                    'level':         p['level'],
                    'quantity':      p['quantity'],
                    'polPortCode':   p['polPortCode'],
                    'podPortCode':   p['podPortCode'],
                    'cargoType':     p['cargoType'],
                    'shipperName':   p['shipperName'],
                    'consigneeName': p['consigneeName'],
                    'confidence':    p['confidence'],
                }
                for p in sol.get('cargoPositions', [])
            ],
        })
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(out, f, indent=2)
    return path

# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    # Load .env.local from project root (one level above stowage-optimizer/)
    env_path = os.path.join(os.path.dirname(__file__), '..', '.env.local')
    load_dotenv(env_path)
    mongo_uri = os.getenv('MONGODB_URI')
    if not mongo_uri:
        print('ERROR: MONGODB_URI not found in .env.local')
        sys.exit(1)

    client = MongoClient(mongo_uri)
    db     = client['reefer-planner']

    # No args → list voyages
    if len(sys.argv) < 2:
        print('Usage: python solver.py <voyage_id>\n')
        print('Available voyages (most recent 20):')
        voyages = list(db.voyages.find(
            {}, {'voyageNumber': 1, 'vesselName': 1, 'status': 1, 'departureDate': 1}
        ).sort('departureDate', -1).limit(20))
        for v in voyages:
            dep = v.get('departureDate', '')
            dep_s = dep.strftime('%Y-%m-%d') if hasattr(dep, 'strftime') else str(dep)[:10]
            print(f"  {v['_id']}  {v.get('voyageNumber','?'):<20}  "
                  f"{v.get('vesselName','?'):<30}  [{v.get('status','?')}]  dep={dep_s}")
        sys.exit(0)

    voyage_id = sys.argv[1]
    print(f'\nLoading voyage {voyage_id}...')

    voyage = load_voyage(db, voyage_id)
    if not voyage:
        print(f'ERROR: Voyage {voyage_id} not found.')
        sys.exit(1)

    vessel = load_vessel(db, voyage.get('vesselId'))
    if not vessel:
        print('ERROR: Vessel not found.')
        sys.exit(1)

    bookings  = load_bookings(db, voyage_id)
    forecasts = load_forecasts(db, voyage_id)

    if not bookings and not forecasts:
        print('WARNING: No confirmed bookings or incorporated forecasts found. Nothing to optimize.')
        sys.exit(0)

    port_seq_map  = build_port_seq_map(voyage.get('portCalls', []))
    sections      = build_sections(vessel)
    cargo_items   = build_cargo_items(bookings, forecasts, port_seq_map)
    voyage_number = voyage.get('voyageNumber', 'UNKNOWN')
    out_dir       = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'output')

    print()
    print_loading_summary(voyage, vessel, bookings, forecasts, sections)

    print(f'\n{"="*62}')
    print(f'CP-SAT optimizer  {len(cargo_items)} cargo items, {len(sections)} sections')
    print(f'{"="*62}\n')

    solutions = []

    for idx, cfg in enumerate(SOLUTION_CONFIGS, start=1):
        print(f'Solving {idx}/5: {cfg["label"]}...')
        assignments, status, solver = solve(cargo_items, sections, cfg, time_limit=30)
        sname = STATUS_NAMES.get(status, str(status))

        if assignments is None:
            print(f'  ✗ {sname}')
            if status == cp_model.INFEASIBLE:
                print('    Likely cause: temperature constraints leave insufficient '
                      'compatible capacity for one or more cargo types, '
                      'or POL/POD monotonicity constraints conflict with '
                      'available section depth ordering.')
            solutions.append({
                'solutionIndex': idx, 'label': cfg['label'],
                'status': sname, 'metrics': {}, 'cargoPositions': [],
            })
            continue

        metrics = compute_metrics(assignments, cargo_items, sections)
        m = metrics
        print(f'  Solution {idx}/5: {cfg["label"]}  [{sname}]')
        print(f'  +-- Pallets placed: {m["placedPallets"]}/{m["totalPallets"]} ({m["placedPct"]:.1f}%)')
        print(f'  +-- OVERSTOW violations: {m["overstowViolations"]}')
        print(f'  +-- Balance score (avg deviation per port): {m["balanceDev"]} pallets')
        print(f'  +-- Compactness score: {m["compactnessPct"]:.1f}% lower levels filled')
        print(f'  +-- Sections used: {m["sectionsUsed"]}/{m["totalSections"]}')
        print()

        cargo_positions = []
        for a in assignments:
            ci  = cargo_items[a['cargoIdx']]
            sec = sections[a['sectionIdx']]
            cargo_positions.append({
                'bookingId':     ci['bookingId'],
                'sectionId':     sec['sectionId'],
                'holdNumber':    sec['holdNumber'],
                'level':         sec['level'],
                'capacity':      sec['capacity'],
                'quantity':      a['quantity'],
                'polPortCode':   ci['polPortCode'],
                'podPortCode':   ci['podPortCode'],
                'cargoType':     ci['cargoType'],
                'shipperName':   ci['shipperName'],
                'consigneeName': ci['consigneeName'],
                'confidence':    ci['confidence'],
            })

        solutions.append({
            'solutionIndex': idx, 'label': cfg['label'],
            'status': sname, 'metrics': metrics,
            'cargoPositions': cargo_positions,
        })

    print(f'{"="*62}')
    print('Exporting results...')
    excel_path = export_excel(solutions, voyage_number, out_dir)
    json_path  = export_json(voyage_id, solutions, voyage_number, out_dir)
    print(f'  Excel : {excel_path}')
    print(f'  JSON  : {json_path}')
    print('\nDone.')


if __name__ == '__main__':
    main()
